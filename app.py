import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import openpyxl


def buscames(conteudo):
    meses = ['Escolha o mês']
    # st.write('buscames')
    for linha in conteudo:
        linha = linha.decode('iso8859-1')
        if linha[9] == "3":
            # print(linha)
            mes = f'{linha[14:18]}-{linha[12:14]}'
            if mes not in meses:
                meses.append(mes)
    meses.sort(reverse=True)
    return meses

def buscaempregado(conteudo):
    empregado = {}
    # print('empregados')
    for linha in conteudo:
        linha = linha.decode('iso8859-1')
        # print(linha)
        if linha[9] == "5":
            # print(linha)
            # print(linha[22])
            if linha[22] == "I" or linha[22] == "A":
                empregado[linha[23:35]] = linha[35:86].strip()
            # if linha[23] == 'E':
            #     del empregado[linha[24:36]]
    return empregado

def buscaponto(conteudo, mes, empregados):
    global df
    datas = []
    horas = []
    funcionarios = []
    for linha in conteudo:
        linha = linha.decode('iso8859-1')
        if linha[9] == "3":
            meslinha = f'{linha[14:18]}-{linha[12:14]}'
            mesant = (datetime.strptime(mes+'-01','%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m')
            if mes == meslinha or mesant  == meslinha:
                dia = datetime.strptime(linha[10:18],"%d%m%Y")
                # print(dia, datetime.strptime(mesant+'-24','%Y-%m-%d'),datetime.strptime(mesbusca+"-26", "%Y-%m-%d"))
                if  dia > datetime.strptime(mesant+'-24','%Y-%m-%d') and dia < datetime.strptime(mes+"-26", "%Y-%m-%d"):
                    # print(mesbusca,mesant,mes,linha[10:18])
                    datas.append(dia)
                    horas.append(f'{linha[18:20]}:{linha[20:22]}')
                    funcionarios.append(empregados[linha[22:34]])
    df = pd.DataFrame(data = {"dia":datas, "hora":horas, "funcionario":funcionarios})

def geraplanilha(funcionario):
    global df, rsocial, cnpj
    fundoescuro = openpyxl.styles.PatternFill("solid", fgColor="111111")
    fonteclara = openpyxl.styles.Font(b=True, color="FFFFFF")
    thin = openpyxl.styles.Side(style='thin',  color="000000")
    borda = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
    dftemp = df[df['funcionario']==funcionario]
    diaant = ""
    coluna = 'A'
    ultimahora = ''
    linhacab = 1
    excel = openpyxl.Workbook()
    planilha = excel.active
    planilha.column_dimensions['A'].width = 15
    planilha.column_dimensions['F'].width = 10

    linha=1
    planilha[f'A{linha}'] = 'Empresa'
    planilha[f'A{linha}'].fill = fundoescuro
    planilha[f'A{linha}'].font = fonteclara
    planilha[f'B{linha}'] =  rsocial
    planilha[f'B{linha}'].border = borda
    planilha.merge_cells(f'B{linha}:F{linha}')
    linha+=1
    planilha[f'A{linha}'] = 'CNPJ'
    planilha[f'A{linha}'].fill = fundoescuro
    planilha[f'A{linha}'].font = fonteclara
    planilha[f'B{linha}'] =  cnpj
    planilha[f'B{linha}'].border = borda
    planilha.merge_cells(f'B{linha}:F{linha}')

    
    linha+=1
    
    planilha[f'A{linha}'] = 'Nome'
    planilha[f'A{linha}'].fill = fundoescuro
    planilha[f'A{linha}'].font = fonteclara
    planilha[f'B{linha}'] =  funcionario
    planilha[f'B{linha}'].border = borda
    planilha.merge_cells(f'B{linha}:F{linha}')

    linha += 2
    linhacab = linha
    planilha[f'A{linha}'] = 'Data'
    planilha[f'B{linha}'] = 'Inicio'
    planilha[f'C{linha}'] = 'Fim'
    planilha[f'D{linha}'] = 'Inicio'
    planilha[f'E{linha}'] = 'Fim'
    planilha[f'F{linha}'] = 'Total'
    for i in range(ord('A'),ord('G')):
        planilha[f'{chr(i)}{linha}'].fill = fundoescuro
        planilha[f'{chr(i)}{linha}'].font = fonteclara

    for i in range(dftemp.shape[0]):
        dia = dftemp.iloc[i]['dia']
        hora = dftemp.iloc[i]['hora']
        if diaant == "":
            coluna = 'A'
            linha+=1
            planilha[f'{coluna}{linha}'] = dia.strftime("%d/%m/%Y")
            coluna = chr(ord(coluna)+1)
            planilha[f'{coluna}{linha}'] = hora
            diaant = dia
            ultimahora = hora
        elif dia != diaant:
            if coluna == 'B' and int(ultimahora[:2]) >17:
                coluna  = chr(ord(coluna)+1)
                planilha[f'{coluna}{linha}'] = "23:59"
            elif coluna == 'B' and int(ultimahora[:2]) < 11:
                planilha[f'{coluna}{linha}'] = "00:00"
                coluna  = chr(ord(coluna)+1)
                planilha[f'{coluna}{linha}'] = ultimahora
            linha+=1
            coluna = 'A'
            planilha[f'{coluna}{linha}'] = dia.strftime("%d/%m/%Y")
            coluna = chr(ord(coluna)+1)
            planilha[f'{coluna}{linha}'] = hora
            diaant = dia
            ultimahora = hora
        else:
            # print("mesma linha")
            if hora != ultimahora:
                coluna = chr(ord(coluna)+1)
                planilha[f'{coluna}{linha}'] = hora
        # print(linha,dia,diaant)
        planilha[f'F{linha}'] = f'=(e{linha} - d{linha})+(c{linha} - b{linha})'
        planilha[f'F{linha}'].number_format = 'HH:MM'
        for i in range(ord('A'),ord('G')):
            planilha[f'{chr(i)}{linha}'].border = borda
            # planilha[f'{chr(i)}{linha}'].font = fonteclara
    linha+=1
    planilha[f'E{linha}'] = 'Total'
    planilha[f'E{linha}'].fill = fundoescuro
    planilha[f'E{linha}'].font = fonteclara
    planilha[f'F{linha}'] = f'=sum(F{linhacab}:F{linha-1})'
    planilha[f'F{linha}'].number_format = '[HH]:MM'
    planilha[f'F{linha}'].border = borda
    
    excel.save(f'{funcionario}.xlsx')




    return dftemp

hide_menu_style = """
        <style>
        #MainMenu {visibility: hidden;}
        </style>
        """

st.set_page_config(
    page_title="processamento de arquivo de ponto",

)
st.markdown(hide_menu_style, unsafe_allow_html=True)

st.title('Processamento de Ponto')
st.write('Processamos apenas arquivos de ponto no formato AFD')
arquivo = st.file_uploader('Escolha um arquivo AFD')
if 'arquivo' in locals() and arquivo is not None:
    if arquivo.type == 'text/plain':
        # st.write("analisar")
        conteudo = arquivo.readlines()
#        st.write(dir(arquivo))
        for linha in conteudo:
            linha = linha.decode('iso8859-1')
            break
        if linha[:9] != '000000000':
            st.error('Arquivo não tem cabeçalho AFD')
        else:
            for linha in conteudo:
                if linha[:9].decode() == "000000000":
                    if linha.decode()[9] == "1":
                        cnpj = linha.decode()[11:25]
                        cnpj = f'{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}'
                        rsocial = linha.decode()[37:187].strip()
                        st.markdown(f'CNPJ: {cnpj}     **{rsocial}**')
                        break
        empregados = buscaempregado(conteudo)
        # mes = datetime.now().strftime("%Y-%m")
        mes = st.selectbox('Escolha o mês: ',(buscames(conteudo)),key="mes")
        # mes = st.selectbox('Escolha o mês: ',(buscames(conteudo)),key="mes")
        # st.write(f'buscando o mes {mes} - > {type(mes)}')
        if mes != "Escolha o mês":
            # st.write(f'buscando o mes {mes} - > {type(mes)}')
            buscaponto(conteudo,mes,empregados)
            for funcionario in df['funcionario'].sort_values().unique():
                geraplanilha(funcionario)
                st.download_button(label=f'{funcionario} - {mes}',file_name=f'{funcionario}-{mes}.xlsx',data=open(f'{funcionario}.xlsx','rb').read())
    else:
        st.error('Este arquivo não é no formato do ponto AFD')